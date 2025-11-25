from scipy.spatial.distance import euclidean
import pandas as pd
import numpy as np
from fastdtw import fastdtw
import openpyxl
import xlwt
import xlrd

filePath01  =r'Table S1. 3,880 processed growth curves.xlsx'
data = pd.read_excel(filePath01)
# datat=data.transpose(copy=False)
book = openpyxl.Workbook()
sheet = book.create_sheet('sheet1')
for i in range(0,3880):
    #print(i)
    made = []
    x = np.array(data.iloc[i])
    # print(x)
    x = x[~np.isnan(x)].flatten()
    h = 3880
    for g in range(i+1,h):
        y = np.array(data.iloc[g])
        y = y[~np.isnan(y)].flatten()
        # print(distance)
        distance, path = fastdtw(x, y, dist=euclidean)
        # print(distance)
        made.append(distance)
    #print(made)
    #print(made)
    for index, a in enumerate(made):
        sheet.cell(2+i, index+2+i, a)
savePath = r'230507第三条DTW0-200.xlsx'
book.save(savePath)

