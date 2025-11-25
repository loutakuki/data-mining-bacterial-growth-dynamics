from __future__ import absolute_import, division
import pandas as pd
import numpy as np
from fastdtw import fastdtw

from scipy.spatial.distance import euclidean
import openpyxl
import xlwt
import xlrd

import numbers
from collections import defaultdict


def euclidean_on_vectors(u, v):
    return np.sqrt(np.sum((u - v) ** 2))

def derivative(x, index):
    # try:
    if len(x) == 0:
        raise Exception("Incorrect input. Must be an array with more than 1 element.")
    elif index == len(x) - 1:
        print("problem")
        return 0
    # print("val", ((x[index] - x[index - 1]) + ((x[index + 1] - x[index - 1])/2))/2)
    return ((x[index] - x[index - 1]) + ((x[index + 1] - x[index - 1]) / 2)) / 2


def derivative_metric(x, y, x_index, y_index):
    # print("inside metrc", x, y, x_index, y_index)
    if x_index == 0 or y_index == 0:
        print("problem")
    elif x_index == len(x) or y_index == len(y):
        print("problem")
    else:
        # print("value", (derivative(x, x_index) - derivative(y, y_index))**2)

        return (derivative(x, x_index) - derivative(y, y_index)) ** 2


def __dtw(x, y, window, dist):
    len_x, len_y = len(x), len(y)
    if window is None:
        window = [(i, j) for i in range(1, len_x - 1) for j in range(1, len_y - 1)]
    window = [(i + 1, j + 1) for i, j in window]
    # print("window", len(window))
    D = defaultdict(lambda: (float('inf'),))
    D[1, 1] = (0, 0, 0)
    for i, j in window:
        '''if i == 0 or j == 0:
            continue
        elif i == len_x - 1 or j == len_y - 1:
            continue'''
        # print("First for loop vals", i - 1, j - 1)
        dt = dist(x, y, i - 1, j - 1)
        D[i, j] = min((D[i - 1, j][0] + dt, i - 1, j), (D[i, j - 1][0] + dt, i, j - 1),
                      (D[i - 1, j - 1][0] + dt, i - 1, j - 1), key=lambda a: a[0])
    path = []
    i, j = len_x - 1, len_y - 1
    while not (i == j == 1):
        '''if i == 0 or j == 0:
            break
        elif i == len_x - 1 or j == len_y - 1:
            continue'''
        # print("indices", (i, j))
        try:
            path.append((i - 1, j - 1))
            i, j = D[i, j][1], D[i, j][2]
        except IndexError:
            print("Getting IndexError here", D[i, j])
    path.reverse()
    return (D[len_x - 1, len_y - 1][0])#, path

# Prepare an Excel workbook and sheet
book = openpyxl.Workbook()
sheet = book.create_sheet('sheet1')

# Load the data
filePath =f'Table S7. 10,247 growth curves.xlsx'
df = pd.read_excel(filePath)

# Prepare a DataFrame for saving selected series
selected_series_df = pd.DataFrame()

# Group by experiment name
experiments = df.groupby(df.columns.str.split('.').str[0], axis=1)

row_counter = 1  # Excel rows start from 1
for experiment_name, experiment_data in experiments:
    series = [data.dropna().to_numpy(dtype='float64') for _, data in experiment_data.iteritems()]
    #print(series)
    #print(len(series))

    if len(series) < 2:  # Skip if only one series for the experiment
        continue
    valid_pairs=[]
    # Compute fastdtw for every pair of series within the same experiment
    for i in range(len(series)):
        for j in range(i + 1, len(series)):
            #distance, _ = fastdtw(series[i], series[j], dist=euclidean_on_vectors)
            distance = __dtw(series[i], series[j], None, derivative_metric)
            sheet.cell(row_counter, 1, f"{experiment_name}_{i + 1}_{j + 1}")
            sheet.cell(row_counter, 2, distance)
            row_counter += 1

            # If the distance is less than 1.1, save this pair to valid_pairs list
            if distance < 0.00053357: #0.00053357:#1.105 DTW25% 0.554DTW 10% #0.0002924DDTW15% #DTW0.554 #0.398
                valid_pairs.append((i, j))

    if valid_pairs:
        unique_indices = list(set(sum(valid_pairs, ())))
        selected_series_df = pd.concat([selected_series_df, experiment_data.iloc[:, unique_indices]], axis=1)

# Save the workbook
savePath = f'DDTW distances of growth curves between experimental replicates.xlsx'
book.save(savePath)

# Save the selected series to a new Excel file
selected_series_df.to_excel(f'selected_seriesby_DDTW_25percentage.xlsx')
