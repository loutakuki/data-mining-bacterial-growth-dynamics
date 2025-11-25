from __future__ import absolute_import, division
from scipy.spatial.distance import euclidean
import pandas as pd
import openpyxl
import xlwt
import xlrd

import numbers
import numpy as np
from collections import defaultdict

def derivative(x, index):
    # try:
    if len(x) == 0:
        raise Exception("Incorrect input. Must be an array with more than 1 element.")
    elif index == len(x) - 1:
        print("problem")
        return 0
    # print("val", ((x[index] - x[index - 1]) + ((x[index + 1] - x[index - 1])/2))/2)
    return ((x[index] - x[index - 1]) + ((x[index + 1] - x[index - 1]) / 2)) / 2


def derivative_metric(x, y, x_index, y_index):
    # print("inside metrc", x, y, x_index, y_index)
    if x_index == 0 or y_index == 0:
        print("problem")
    elif x_index == len(x) or y_index == len(y):
        print("problem")
    else:
        # print("value", (derivative(x, x_index) - derivative(y, y_index))**2)

        return (derivative(x, x_index) - derivative(y, y_index)) ** 2


def __dtw(x, y, window, dist):
    len_x, len_y = len(x), len(y)
    if window is None:
        window = [(i, j) for i in range(1, len_x - 1) for j in range(1, len_y - 1)]
    window = [(i + 1, j + 1) for i, j in window]
    # print("window", len(window))
    D = defaultdict(lambda: (float('inf'),))
    D[1, 1] = (0, 0, 0)
    for i, j in window:
        '''if i == 0 or j == 0:
            continue
        elif i == len_x - 1 or j == len_y - 1:
            continue'''
        # print("First for loop vals", i - 1, j - 1)
        dt = dist(x, y, i - 1, j - 1)
        D[i, j] = min((D[i - 1, j][0] + dt, i - 1, j), (D[i, j - 1][0] + dt, i, j - 1),
                      (D[i - 1, j - 1][0] + dt, i - 1, j - 1), key=lambda a: a[0])
    path = []
    i, j = len_x - 1, len_y - 1
    while not (i == j == 1):
        '''if i == 0 or j == 0:
            break
        elif i == len_x - 1 or j == len_y - 1:
            continue'''
        # print("indices", (i, j))
        try:
            path.append((i - 1, j - 1))
            i, j = D[i, j][1], D[i, j][2]
        except IndexError:
            print("Getting IndexError here", D[i, j])
    path.reverse()
    return (D[len_x - 1, len_y - 1][0])#, path

filePath01 = r'Table S1. 3,880 processed growth curves.xlsx'
data = pd.read_excel(filePath01)
# datat=data.transpose(copy=False)
book = openpyxl.Workbook()
sheet = book.create_sheet('sheet1')
#num=4
for i in range(0,3880):#(0,2)+(2,4)组成（0,4）..range(4)=range(0,4)
    #print(i)
    made = []
    x = np.array(data.iloc[i])
    # print(x)
    x = x[~np.isnan(x)]
    h = 3880
    for g in range(i+1,h):
        y = np.array(data.iloc[g])
        y = y[~np.isnan(y)]
        # print(distance)
        made.append(__dtw(x,y,None,derivative_metric))
    #print(made)
    #print(made)
    for index, a in enumerate(made):
        sheet.cell(2+i, index+2+i, a)#(行，列，数值)

savePath = r'Table S3. DDTW matrix of 3,880 growth curves.xlxs'
book.save(savePath)


